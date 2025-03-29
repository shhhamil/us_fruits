from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse
from django.utils import timezone
from django.http import Http404
from django.contrib.auth import authenticate, login, logout
from django.db.models import Count
from user_profile.models import Complaint , Wallet
from django.utils.dateparse import parse_datetime
from django.contrib import messages
from decimal import Decimal
from .models import Coupon
from django.views.decorators.cache import never_cache
from django.contrib.auth.decorators import login_required
from django.utils.dateparse import parse_date
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from .models import Category, Product,OrderItem
from .models import CustomUser
from django.contrib import messages
from .models import Order
from django.db.models import Sum
from datetime import datetime, timedelta
from django.db.models import Q
from django.views.decorators.http import require_POST
import json
from django.db.models import Sum
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill  
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from io import BytesIO
# ______________________________________________________________________________________________________________
# ---------------------------------------LOGIN VIEWS------------------------------------------------------------
def admin_login(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        password = request.POST.get('password')
        if not email or not password:
            messages.error(request, 'Email and password are required.')
            return redirect('Admin_Login')
        user = authenticate(request, username=email, password=password)
        if user is not None:
            if user.is_superuser:
                if not user.is_active:
                    messages.error(request, "You're blocked. Please contact support.")
                    return redirect('Admin_Login')
                login(request, user)
                return redirect('Admin_Dashbord')
            else:
                messages.error(request, 'You do not have permission to access the admin panel.')
        else:
            messages.error(request, 'Invalid user credentials')
        return redirect('Admin_Login')
    return render(request, 'auth_admin/Admin_login.html')
# ______________________________________________________________________________________________________________
# ------------------------------------DASHBORD VIEWWS-----------------------------------------------------------
@never_cache
def admin_dashbord(request):
    if not request.user.is_superuser:
        messages.error(request, 'You are not authorized to access this page.')
        return redirect('Admin_Login')
    
    popular_products = Product.objects.annotate(order_count=Count('orderitem')).order_by('-order_count')[:5]
    popular_categories = Category.objects.annotate(order_count=Count('products__orderitem')).order_by('-order_count')[:5]
    

    
    context = {
        'popular_products': popular_products,
        'popular_categories': popular_categories,
    }
    return render(request, 'auth_admin/admin_dashbord.html', context)
# ______________________________________________________________________________________________________________
# ---------------------------------------LOGOUT VIEWS-----------------------------------------------------------
@never_cache
def admin_logout(request):
    logout(request)   
    request.session.flush() 
    return redirect('Admin_Login')
# ______________________________________________________________________________________________________________
# -----------------------------------------View for category management page------------------------------------
@login_required
def Category_Manegement(request):
    page = request.GET.get('page', 1)   
    search = request.GET.get('search', '')   
    categories = Category.objects.all()
    
    if search:
        categories = categories.filter(name__icontains=search)
    
    paginator = Paginator(categories, 10)  
    try:
        categories_page = paginator.page(page)
    except PageNotAnInteger:
        categories_page = paginator.page(1)  
    except EmptyPage:
        categories_page = paginator.page(paginator.num_pages)   
    
    context = {'categories': categories_page, 'total_categories': categories.count(), 'search_term': search}
    return render(request, 'auth_admin/category.html', context)
# _______________________________________________________________________________________________________________
# ================================= Add Category =============================================================
def add_category(request):
    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        discount = request.POST.get('discount', '').strip()

        if not name:
            messages.error(request, "Category name is required!")
            return redirect('Add_Category')

        if not description:
            messages.error(request, "Description is required!")
            return redirect('Add_Category')

        if discount:
            try:
                discount = Decimal(discount) 
                if discount < 0 or discount > 100:
                    messages.error(request, "Discount must be between 0 and 100!")
                    return redirect('Add_Category')
            except ValueError:
                messages.error(request, "Invalid discount value!")
                return redirect('Add_Category')
        else:
            discount = Decimal(0)  
        if Category.objects.filter(name__iexact=name).exists():
            messages.error(request, "Category already exists!")
            return redirect('Add_Category')

        Category.objects.create(
            name=name,
            description=description,
            discount_percentage=discount
        )
        messages.success(request, "Category added successfully!")
        return redirect('Category')

    categories = Category.objects.all()  
    return render(request, 'auth_admin/add_cate.html', {"categories": categories})
# ============================== Edit Category ===============================================================
def edit_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if request.method == "POST":
        name = request.POST.get('name', '').strip()
        description = request.POST.get('description', '').strip()
        discount = request.POST.get('discount', '').strip()

        if not name:
            messages.error(request, "Category name is required!")
            return redirect('Edit_Category', category_id=category.id)

        if not description:
            messages.error(request, "Description is required!")
            return redirect('Edit_Category', category_id=category.id)

        if discount:
            try:
                discount = Decimal(discount) 
                if discount < 0 or discount > 100:
                    messages.error(request, "Discount must be between 0 and 100!")
                    return redirect('Edit_Category', category_id=category.id)
            except ValueError:
                messages.error(request, "Invalid discount value!")
                return redirect('Edit_Category', category_id=category.id)
        else:
            discount = category.discount_percentage   

        if Category.objects.filter(name__iexact=name).exclude(id=category.id).exists():
            messages.error(request, "Category with this name already exists!")
            return redirect('Edit_Category', category_id=category.id)

        category.name = name
        category.description = description
        category.discount_percentage = discount
        category.save()

        messages.success(request, "Category updated successfully!")
        return redirect('Category')

    return render(request, 'auth_admin/edit_cate.html', {'category': category})
# --------------------------------------Delete Category---------------------------------------------------------
def soft_delete_category(request, category_id):
    category = get_object_or_404(Category, id=category_id)

    if category.is_active:
        category.is_active = False
        messages.success(request, "Category disabled successfully!")
    else:
        category.is_active = True
        messages.success(request, "Category restored successfully!")

    category.save()
    return redirect('Category')
# _______________________________________________________________________________________________________________
# ----------------------------------------PRODUCT VIEWS----------------------------------------------------------
@login_required
def Product_Management(request):
    page = request.GET.get('page', 1)
    search = request.GET.get('search', '')
    
    products = Product.objects.filter(is_active=True).order_by('name')
    if search:
        products = products.filter(name__icontains=search)
    
    paginator = Paginator(products, 5)
    try:
        products_page = paginator.page(page)
    except PageNotAnInteger:
        products_page = paginator.page(1)
    except EmptyPage:
        products_page = paginator.page(paginator.num_pages)
    
    context = {
        'products': products_page,
        'total_products': products.count(),
        'search_term': search
    }
    return render(request, 'auth_admin/product.html', context)
# -------------------------------------View to add a new product-------------------------------------------------
def add_product(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category')
        price = request.POST.get('price')
        offer_percentage = request.POST.get('offer_percentage', 0)
        stock = request.POST.get('stock')
        is_active = request.POST.get('is_active') == 'on'
        description = request.POST.get('description', '').strip()
        photo_1 = request.FILES.get('photo_1')
        photo_2 = request.FILES.get('photo_2')
        photo_3 = request.FILES.get('photo_3')

        try:
            price = Decimal(price)  
            stock = int(stock)
            offer_percentage = Decimal(offer_percentage)   
        except ValueError:
            messages.error(request, "Invalid price, stock, or discount value!")
            return redirect('add_product')

        if not name:
            messages.error(request, "Product name is required!")
            return redirect('add_product')
        if not description:
            messages.error(request, "Description is required!")
            return redirect('add_product')
        if not category_id:
            messages.error(request, "Category is required!")
            return redirect('add_product')
        if price < 0:
            messages.error(request, "Price cannot be negative!")
            return redirect('add_product')
        if not (0 <= offer_percentage <= 100):
            messages.error(request, "Offer percentage must be between 0 and 100!")
            return redirect('add_product')
        if stock < 0:
            messages.error(request, "Stock cannot be negative!")
            return redirect('add_product')
        if not photo_1:
            messages.error(request, "At least one photo is required!")
            return redirect('add_product')

        category = get_object_or_404(Category, id=category_id)

        offer_price = price - (price * (offer_percentage / Decimal(100)))
        offer_price = offer_price.quantize(Decimal('0.01'))  

        try:
            Product.objects.create(
                name=name,
                category=category,
                price=price,
                offer_percentage=offer_percentage,
                offer_price=offer_price,
                stock=stock,
                is_active=is_active,
                description=description,
                photo_1=photo_1,
                photo_2=photo_2,
                photo_3=photo_3
            )
            messages.success(request, "Product added successfully!")
            return redirect('Product')
        except Exception as e:
            messages.error(request, f"Error creating product: {str(e)}")
            return redirect('add_product')

    categories = Category.objects.all()
    return render(request, 'auth_admin/add_pro.html', {'categories': categories})
# ---------------- ---------------------------Edit Product View ------------------------------------------------
def edit_product(request, pk):
    product = get_object_or_404(Product, id=pk)
    categories = Category.objects.all()

    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        category_id = request.POST.get('category')
        price = request.POST.get('price')
        offer_percentage = request.POST.get('offer_percentage', 0)
        stock = request.POST.get('stock')
        is_active = request.POST.get('is_active') == 'on'
        description = request.POST.get('description', '').strip()
        photo_1 = request.FILES.get('photo_1')
        photo_2 = request.FILES.get('photo_2')
        photo_3 = request.FILES.get('photo_3')

        try:
            price = Decimal(price) 
            stock = int(stock)
            offer_percentage = Decimal(offer_percentage)   
        except ValueError:
            messages.error(request, "Invalid price, stock, or discount value!")
            return redirect('Edit_Product', pk=pk)

        if not name:
            messages.error(request, "Product name is required!")
            return redirect('Edit_Product', pk=pk)
        if not description:
            messages.error(request, "Description is required!")
            return redirect('Edit_Product', pk=pk)
        if not category_id:
            messages.error(request, "Category is required!")
            return redirect('Edit_Product', pk=pk)
        if price < 0:
            messages.error(request, "Price cannot be negative!")
            return redirect('Edit_Product', pk=pk)
        if not (0 <= offer_percentage <= 100):
            messages.error(request, "Offer percentage must be between 0 and 100!")
            return redirect('Edit_Product', pk=pk)
        if stock < 0:
            messages.error(request, "Stock cannot be negative!")
            return redirect('Edit_Product', pk=pk)

        category = get_object_or_404(Category, id=category_id)

        try:
            product.name = name
            product.category = category
            product.price = price
            product.offer_percentage = offer_percentage
            product.stock = stock
            product.is_active = is_active
            product.description = description

           
            product.offer_price = price - (price * (offer_percentage / Decimal(100)))
            product.offer_price = product.offer_price.quantize(Decimal('0.01'))

            if photo_1:
                product.photo_1 = photo_1
            if photo_2:
                product.photo_2 = photo_2
            if photo_3:
                product.photo_3 = photo_3

            product.save()
            messages.success(request, "Product updated successfully!")
            return redirect('Product')

        except Exception as e:
            messages.error(request, f"Error updating product: {str(e)}")
            return redirect('Edit_Product', pk=pk)

    return render(request, 'auth_admin/edit_pro.html', {'product': product, 'categories': categories})
# ________________________________________________________________________________________________________________________________
#------------------------------------------ View to see the details of a product---------------------------------------------------
def view_product(request, pk):
    product = get_object_or_404(Product, id=pk)
    images = [img.url for img in [product.photo_1, product.photo_2, product.photo_3] if img]
    
  
    final_price = product.offer_price if product.offer_percentage > 0 else product.price

    return render(request, 'auth_admin/view_pro.html', {
        'product': product,
        'images': images,
        'final_price': final_price,  
    })
# -------------------------View to soft delete a product--------------------------------------------------------------------------------
def soft_delete_product(request, pk):
    product = get_object_or_404(Product, id=pk)
    product.is_active = not product.is_active
    product.save()
    if product.is_active:
        messages.success(request, "Product restored successfully!")
    else:
        messages.success(request, "Product disabled successfully!")
    return redirect('Product')
# ______________________________________________________________________________________________________________________________
# -------------------------------------------------User Managment --------------------------------------------------------------------------
login_required
def user_management(request):
    page = request.GET.get('page', 1)  
    search = request.GET.get('search', '').strip()
    users = CustomUser.objects.filter(Q(username__icontains=search) | Q(email__icontains=search),is_active=True).order_by('id')

    paginator = Paginator(users, 7)  
    try:
        users_page = paginator.page(page)
    except (PageNotAnInteger, ValueError):  
        users_page = paginator.page(1)   
    except EmptyPage:
        users_page = paginator.page(paginator.num_pages)  

    context = {
        'users': users_page,
        'total_users': users.count(),
        'search_term': search  
    }
    return render(request, 'auth_admin/user.html', context)
# -------------------------------------------Block User------------------------------------------------------------------------------
def block_user(request, user_id):
    user = get_object_or_404(CustomUser, id=user_id)

    if user.is_superuser:
        messages.error(request, "You cannot block an admin user!")
        return redirect("User_Management")

    user.is_active = not user.is_active
    user.save()

    status = "unblocked" if user.is_active else "blocked"
    messages.success(request, f"User {user.username} has been {status}.")
    
    return redirect("User_Management")
# ---------------------------------------------------------------------------------------------------------------------------------
# _________________________________________Order Views______________________________________________________________________
@login_required
def order_management(request):
    status_filter = request.GET.get('status', 'all').strip().lower()
    orders = Order.objects.all().order_by('-date_of_order')

    if status_filter in dict(Order.STATUS_CHOICES): 
        orders = orders.filter(status=status_filter)

    paginator = Paginator(orders, 5)
    page = request.GET.get('page', 1)

    try:
        orders_page = paginator.page(page)
    except PageNotAnInteger:
        orders_page = paginator.page(1) 
    except EmptyPage:
        orders_page = paginator.page(paginator.num_pages)  

    total_orders = paginator.count  
    total_price = orders.aggregate(Sum('total'))['total__sum'] or 0  


    complaints = Complaint.objects.filter(order__in=orders, order__status='delivered')
    complaints_dict = {complaint.order.id: complaint for complaint in complaints}

    for order in orders_page:
        order.complaint = complaints_dict.get(order.id)   

    context = {
        'orders': orders_page,
        'current_status': status_filter,
        'total_orders': total_orders,  
        'total_price': total_price,  
    }
    return render(request, 'auth_admin/order.html', context)

# -----------------------------------------------------COMPLAINT SESSION------------------------------------------------------------------
def admin_view_complaint(request, complaint_id):
    try:
        complaint = Complaint.objects.get(id=complaint_id)
    except Complaint.DoesNotExist:
        raise Http404("Complaint not found")
    
    return render(request, 'auth_admin/admin_comp.html', {'complaint': complaint})
# ---------------------------------------------------------------------------------------------------
@login_required
def approve_complaint(request, complaint_id):
    try:
        complaint = Complaint.objects.get(id=complaint_id)
        order = complaint.order
        user = order.user
        refund_amount = order.total  


        complaint.status = 'approved'
        complaint.save()

        
        wallet, created = Wallet.objects.get_or_create(user=user)
        wallet.balance += refund_amount
        wallet.save()

        messages.success(request, f"Refund of ₹{refund_amount} added to {user.email}'s wallet.")
    except Complaint.DoesNotExist:
        messages.error(request, "Complaint not found.")
    
    return redirect('Order_Management')
# ----------------------------------------------------------------------------------------
@login_required
def reject_complaint(request, complaint_id):
    try:
        complaint = Complaint.objects.get(id=complaint_id)
        complaint.status = 'rejected'
        complaint.save()

        messages.warning(request, "Complaint has been rejected.")
    except Complaint.DoesNotExist:
        messages.error(request, "Complaint not found.")
    
    return redirect('Order_Management')

# __________________________________________________________________________________________________________________________________
# ----------------------------------------------------------------------------------------------------------------------------------
@login_required
@require_POST
def update_order_status(request, orderId):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'User not authenticated!'}, status=401)

    try:
        data = json.loads(request.body)
        new_status = data.get('status', '').strip().lower()

        order = get_object_or_404(Order, id=orderId)
        valid_statuses = ["pending", "shipped", "delivered", "cancelled"]

        if new_status not in valid_statuses:
            return JsonResponse({'success': False, 'message': 'Invalid order status!'}, status=400)

        order.status = new_status
        if new_status == "delivered":
            order.payment_status = "paid"

        order.save()

        return JsonResponse({
            'success': True,
            'order_status': order.status,
            'payment_status': order.payment_status
        })

    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'message': 'Invalid JSON data!'}, status=400)

    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)}, status=500)
#  -----------------------------this for getting all user detials------------------------------------------------------
def get_order_details(request, order_id):
    if request.user.is_superuser:
        order = get_object_or_404(Order, id=order_id)
    else:
        order = get_object_or_404(Order, id=order_id, user=request.user)

    return render(request, 'auth_admin/admin_order_view.html', {'order': order})


# --------------------------------------sales report---------------------------------------------------------------
@login_required
def sales_report(request):
    period = request.GET.get('period', 'all')
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    page = request.GET.get('page', 1)
    today = timezone.now().date()
    
    if period == 'today':
        orders = Order.objects.filter(date_of_order__date=today)
    elif period == 'weekly':
        start_of_week = today - timedelta(days=today.weekday())
        orders = Order.objects.filter(date_of_order__date__gte=start_of_week)
    elif period == 'monthly':
        start_of_month = today.replace(day=1)
        orders = Order.objects.filter(date_of_order__date__gte=start_of_month)
    elif period == 'yearly':
        start_of_year = today.replace(month=1, day=1)
        orders = Order.objects.filter(date_of_order__date__gte=start_of_year)
    elif period == 'custom' and start_date and end_date:
        try:
            start_date_obj = parse_date(start_date)
            end_date_obj = parse_date(end_date)
            if end_date_obj < start_date_obj:
                return render(request, 'auth_admin/sales.html', {'error': 'End date must be after start date','period': period})
            orders = Order.objects.filter(date_of_order__date__range=[start_date_obj, end_date_obj])
        except (ValueError, TypeError):
            return render(request, 'auth_admin/sales.html', {'error': 'Invalid date format','period': period})
    else:
        orders = Order.objects.all()
    
    orders = orders.order_by('-date_of_order')
    
    total_sales = orders.aggregate(total=Sum('total'))['total'] or 0
    total_orders = orders.count()
    total_products_sold = OrderItem.objects.filter(order__in=orders).aggregate(total=Sum('quantity'))['total'] or 0
    completed_orders = orders.filter(status='delivered').count()
    cancelled_orders = orders.filter(status='cancelled').count()   
    
    stats = {
        'total_sales': total_sales,
        'total_orders': total_orders,
        'total_products': total_products_sold,
        'completed_orders': completed_orders,
        'cancelled_orders': cancelled_orders,
    }
    paginator = Paginator(orders, 10)
    try:
        paginated_orders = paginator.page(int(page))
    except (PageNotAnInteger, EmptyPage):
        paginated_orders = paginator.page(1)
    context = {
        'orders': paginated_orders,
        'stats': stats,
        'period': period,
        'message': request.GET.get('message')
    }
    
    return render(request, 'auth_admin/sales.html', context)
# ---------------------------------------------------pdf of sales report-------------------------------------------------------
@login_required
def export_sales_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d")}.pdf"'
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, spaceAfter=30, alignment=1, textColor=colors.HexColor('#82ae46'))
    elements.append(Paragraph("Sales Report", title_style))
    elements.append(Spacer(1, 20))
    orders = Order.objects.all().order_by('-date_of_order')
    table_data = [['Order ID', 'Date', 'Customer', 'Total Amount', 'Status', 'Payment']]
    for order in orders:
        row = [
            f'#{order.id}',
            order.date_of_order.strftime('%Y-%m-%d'),
            order.user.username if order.user else 'Guest',
            f'₹{order.total:,.2f}',
            order.status,
            order.payment_status
        ]
        table_data.append(row)
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#82ae46')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response
# ------------------------------------------exel sheet of sales report------------------------------------------
@login_required
def export_sales_excel(request):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    headers = ['Order ID', 'Date', 'Customer', 'Total Amount', 'Status', 'Payment']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='82AE46', fill_type='solid')
    orders = Order.objects.all().order_by('-date_of_order')
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=f'#{order.id}')
        ws.cell(row=row, column=2, value=order.date_of_order.strftime('%Y-%m-%d'))
        ws.cell(row=row, column=3, value=order.user.username if order.user else 'Guest')
        ws.cell(row=row, column=4, value=f'₹{order.total:.2f}')
        ws.cell(row=row, column=5, value=order.status)
        ws.cell(row=row, column=6, value=order.payment_status)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{datetime.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response
# ---------------------------------------filter recored ------------------------------------------------------
@login_required
def filter_sales(request, period):
    today = datetime.today()
    if period == 'daily':
        start_date = today.replace(hour=0, minute=0, second=0)
        end_date = today.replace(hour=23, minute=59, second=59)
    elif period == 'weekly':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif period == 'monthly':
        start_date = today.replace(day=1)
        end_date = today.replace(day=28) + timedelta(days=4)
    elif period == 'yearly':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    else:
        return render(request, 'sales/dashboard.html', {'error': 'Invalid period'})
    orders = Order.objects.filter(date_of_order__range=[start_date, end_date])
    return render(request, 'auth_admin/sales.html', {'orders': orders})
# -------------------------------custom date filtering------------------------------------------------------
@login_required
def filter_custom_date(request):
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    if start_date and end_date:
        start_date = datetime.strptime(start_date, "%Y-%m-%d").date()
        end_date = datetime.strptime(end_date, "%Y-%m-%d").date()
        orders = Order.objects.filter(date_of_order__range=(start_date, end_date))
    else:
        orders = Order.objects.none()
    return render(request, 'auth_admin/sales.html', {'orders': orders})

# --------------------------------COUPON MANAGMENT -------------------------------------------------------------
def coupon_list(request):
    coupons = Coupon.objects.all()
    return render(request, "auth_admin/coupon.html", {"coupons": coupons})
# ---------------------------------------------------------------------------------------------------------
def add_coupon(request):
    if request.method == "POST":
        code = request.POST.get("code", "").strip()
        discount_amount = request.POST.get("discount_amount", "0")   
        min_order_amount = request.POST.get("minimum_order_amount", "0")   
        usage_limit = request.POST.get("usage_limit", "1")
        valid_from = request.POST.get("start_date", "")  
        valid_to = request.POST.get("end_date", "")  

        if not code or not valid_from or not valid_to:
            return JsonResponse({"success": False, "error": "Missing required fields"}, status=400)

        try:
            discount_amount = float(discount_amount)
            min_order_amount = float(min_order_amount)
            usage_limit = int(usage_limit)
            valid_from = parse_datetime(valid_from)
            valid_to = parse_datetime(valid_to)

            if valid_from is None or valid_to is None:
                return JsonResponse({"success": False, "error": "Invalid date format"}, status=400)

            coupon = Coupon.objects.create(
                code=code,
                discount_amount=discount_amount,
                min_order_amount=min_order_amount,
                usage_limit=usage_limit,
                start_date=valid_from,
                end_date=valid_to,
            )

            return JsonResponse({"success": True})

        except ValueError as e:
            return JsonResponse({"success": False, "error": "Invalid data type"}, status=400)

    return JsonResponse({"success": False}, status=400)
# ---------------------------------------------------------------------------------
def edit_coupon(request, coupon_id):
    if request.method == "POST":
        coupon = get_object_or_404(Coupon, id=coupon_id)
        code = request.POST.get("code")
        discount_amount = request.POST.get("discount_amount")
        min_order_amount = request.POST.get("min_order_amount")
        usage_limit = request.POST.get("usage_limit")
        start_date = request.POST.get("start_date")
        end_date = request.POST.get("end_date")

        try:
            if not min_order_amount:
                return JsonResponse({"success": False, "error": "Minimum order amount is required."})

            discount_amount = float(discount_amount)  
            min_order_amount = Decimal(min_order_amount)   

            if discount_amount > min_order_amount:
                return JsonResponse({"success": False, "error": "Discount amount cannot be greater than minimum order amount."})

        except ValueError:
            return JsonResponse({"success": False, "error": "Discount amount and minimum order amount must be valid numbers."})

        if usage_limit:
            try:
                usage_limit = int(usage_limit)
            except ValueError:
                return JsonResponse({"success": False, "error": "Usage limit must be a valid integer."})


        start_date = parse_datetime(start_date)
        end_date = parse_datetime(end_date)

        if not start_date or not end_date:
            return JsonResponse({"success": False, "error": "Invalid date format."})

        if start_date >= end_date:
            return JsonResponse({"success": False, "error": "Start date must be earlier than end date."})

        coupon.code = code
        coupon.discount_amount = discount_amount
        coupon.min_order_amount = min_order_amount
        coupon.usage_limit = usage_limit
        coupon.start_date = start_date
        coupon.end_date = end_date
        coupon.save()

        return JsonResponse({"success": True, "message": "Coupon updated successfully!"})

    return JsonResponse({"success": False, "error": "Invalid request method."}, status=405)
# --------------------------------------------------------------------------------
def delete_coupon(request, coupon_id):
    coupon = Coupon.objects.get(id=coupon_id)
    coupon.toggle_delete()
    return redirect("coupon_list")